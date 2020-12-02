# coding=utf-8
import openpyxl
import pandas as pd


# 判断该列的数据类型
def column_type(s):
    if s == 'ID':
        # 数据为int类型
        return 1
    else:
        if s == 'Name' or s == 'City' or s == 'Gender' or s == 'Constitution\n':
            # 数据为str类型
            return 2
        else:
            # 数据为float类型
            return 3


def txt2xlsx(filename, output_filename):
    with open(filename, 'r', encoding="utf-8") as f:
        # 循环读取文本文件的每一行数据
        col = 1
        # 列名
        columnsname = []
        # 储存文本数据的字典
        dict_data = {}
        while col:
            line = f.readline()  # 一行一行读取
            if not line:  # 如果没有内容，则退出循环
                break
            # 文本文件的第一行是columns name
            if col == 1:
                for i in line.split(','):  # 将每一行的数据进行切片处理
                    columnsname.append(i)
            else:
                # 用来遍历columnsname列表的变量
                c = 0
                for i in line.split(','):
                    # 去掉回车符'\n'
                    if '\n' in i:
                        i = i.replace('\n', '')
                    # columnsname[c]是键名
                    # 对数据进行类型转换
                    if column_type(columnsname[c]) == 1 and i != '':
                        i = int(i)
                    else:
                        if column_type(columnsname[c]) == 3 and i != '':
                            i.strip()
                            i = float(i)
                    dict_data.setdefault(columnsname[c], []).append(i)
                    c = c + 1
            col = col + 1
    frame = pd.DataFrame(dict_data, columns=columnsname)
    frame.to_excel(output_filename, sheet_name='sheet1', na_rep='', float_format=None, columns=None,
                   header=True, index=False, index_label=None, startrow=0, startcol=0, engine=None,
                   merge_cells=True, encoding=None, inf_rep='inf', verbose=True, freeze_panes=None)
    print("转换完成")


# 规范学号、身高、性别
def standard(filename1, filename2):
    # 打开excel，并获取最大行数
    wb1 = openpyxl.load_workbook(filename1)
    ws1 = wb1.active
    row1 = ws1.max_row
    wb2 = openpyxl.load_workbook(filename2)
    ws2 = wb2.active
    row2 = ws2.max_row
    # 规范file1
    for r in range(2, row1 + 1):
        # 规范学号
        if ws1.cell(row=r, column=1).value is not None:
            if int(ws1.cell(row=r, column=1).value) < 202000:
                ws1.cell(row=r, column=1).value = int(ws1.cell(row=r, column=1).value) + 202000
        # 规范性别
        if ws1.cell(row=r, column=4).value is not None:
            if ws1.cell(row=r, column=4).value == 'boy':
                ws1.cell(row=r, column=4).value = 'male'
            else:
                if ws1.cell(row=r, column=4).value == 'girl':
                    ws1.cell(row=r, column=4).value = 'female'
        # 规范身高
        if ws1.cell(row=r, column=5).value is not None:
            if int(ws1.cell(row=r, column=5).value) > 100:
                ws1.cell(row=r, column=5).value = int(ws1.cell(row=r, column=5).value) / 100
    # 规范file2
    for r in range(2, row2 + 1):
        # 规范学号
        if ws2.cell(row=r, column=1).value is not None:
            if int(ws2.cell(row=r, column=1).value) < 202000:
                ws2.cell(row=r, column=1).value = int(ws2.cell(row=r, column=1).value) + 202000
        # 规范性别
        if ws2.cell(row=r, column=4).value is not None:
            if ws2.cell(row=r, column=4).value == 'boy':
                ws2.cell(row=r, column=4).value = 'male'
            else:
                if ws2.cell(row=r, column=4).value == 'girl':
                    ws2.cell(row=r, column=4).value = 'female'
        # 规范身高
        if ws2.cell(row=r, column=5).value is not None:
            if int(ws2.cell(row=r, column=5).value) > 100:
                ws2.cell(row=r, column=5).value = int(ws2.cell(row=r, column=5).value) / 100
    wb1.save(filename1)
    wb2.save(filename2)
    print("规范完成")


# 读取excel整行的数据
def read_excle_line(n, filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    row = ws[n]
    line = list()
    for cell in row:
        line.append(cell.value)
    return line


# 计算缺失度，空缺数据的位置
def missing_degree(a):
    m = 0
    empty = list()
    for i in range(len(a)):
        if a[i] is None:
            if i != 14:
                m = m + 1
                empty.append(i)
    return m, empty


# 删除名字的特殊字符
def delete_special(s):
    special = ['?', '/', '.', '@', '#', '$', '%', '^']
    for i in range(len(special)):
        if special[i] in s:
            # 删除特殊字符
            s = s.replace(special[i], '')
            return s
    return None


# 数据择优函数
def select(a, b):
    # 计算缺失度
    ma, empty_a = missing_degree(a)
    mb, empty_b = missing_degree(b)
    # 选择缺失度少的数据
    if ma <= mb:
        return 1, empty_a
    else:
        return 2, empty_b


# 根据学号查名字
def check_name(n, filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    row = ws.max_row
    for r in range(2, row + 1):
        if n == ws.cell(row=r, column=1).value:
            if n != ws.cell(row=r + 1, column=1).value:
                return ws.cell(row=r, column=2).value
            else:
                if ws.cell(row=r, column=1).value == ws.cell(row=r + 1, column=1).value:
                    return ws.cell(row=r, column=2).value
                else:
                    return None
    return None


# 去除冗余函数
def remove(filename_1):
    # 记录空缺值的位置，填补空白值用
    empty_position = []
    wb = openpyxl.load_workbook(filename_1)
    ws = wb.active
    r = 2
    while ws.cell(row=r, column=1).value is not None:
        if ws.cell(row=r, column=1).value != ws.cell(row=r + 1, column=1).value:
            # 记录空白值
            a = read_excle_line(r, filename_1)
            m, empty = missing_degree(a)
            for i in range(len(empty)):
                position = (r, empty[i] + 1)
                empty_position.append(position)
            # 没有出现冗余，r + 1
            r = r + 1
        else:
            # 如果出现冗余
            a = read_excle_line(r, filename_1)
            b = read_excle_line(r + 1, filename_1)
            # 将r行数据与r+1行数据进行缺失度比较
            stay, empty = select(a, b)
            # 留下第r行
            if stay == 1:
                # 检查第r行数据的名字有无特殊符号
                clearname = delete_special(ws.cell(row=r, column=2).value)
                if clearname is not None:
                    ws.cell(row=r, column=2).value = clearname
                # 根据第r+1行的数据，补充第r行的空缺
                for i in range(len(empty)):
                    if b[empty[i]] is not None:
                        # 填补空缺
                        ws.cell(row=r, column=empty[i] + 1).value = ws.cell(row=r + 1, column=empty[i] + 1).value
                # 删除r+1行,留下r行
                ws.delete_rows(r + 1)
            else:
                # 留下第r+1行
                if stay == 2:
                    # 检查第r+1行数据的名字有无特殊符号
                    clearname = delete_special(ws.cell(row=r + 1, column=2).value)
                    if clearname is not None:
                        ws.cell(row=r + 1, column=2).value = clearname
                    # 根据第r行的数据，补充第r+1行的空缺
                    for i in range(len(empty)):
                        if a[empty[i]] is not None:
                            ws.cell(row=r + 1, column=empty[i] + 1).value = ws.cell(row=r, column=empty[i] + 1).value
                    # 删除r行，留下r+1行
                    ws.delete_rows(r)
            wb.save(filename_1)
    print("去冗余完成")
    return empty_position


# 整合两个数据源的数据
def combine(data_1name, data_2name):
    # 打开excel文件
    wb_1 = openpyxl.load_workbook(data_1name)
    wb_2 = openpyxl.load_workbook(data_2name)
    wb_3 = openpyxl.Workbook()
    # 获取当前活跃表单
    ws_1 = wb_1.active
    ws_2 = wb_2.active
    ws_3 = wb_3.active
    # 获取数据最大行数和列数
    row_1 = ws_1.max_row
    col_1 = ws_1.max_column
    row_2 = ws_2.max_row
    col_2 = ws_2.max_column
    # 将文件1的数据写入新文件中
    print("data_1 starting")
    for r1 in range(1, row_1 + 1):
        for c1 in range(1, col_1 + 1):
            ws_3.cell(row=r1, column=c1).value = ws_1.cell(row=r1, column=c1).value
    # 将文件2的数据写入新文件中
    print("data_2 starting")
    for r2 in range(row_1 + 1, row_1 + row_2 + 1):
        for c2 in range(1, col_2 + 1):
            ws_3.cell(row=r2, column=c2).value = ws_2.cell(row=r2 - row_1 + 1, column=c2).value
    # 保存文件
    wb_3.save('new.xlsx')
    # 对新文件按ID号进行排序
    df = pd.read_excel('new.xlsx')
    df = df.sort_values('ID', ascending=True)
    # 对完全一样的数据进行去冗余处理
    df = df.drop_duplicates(subset=None, keep='first', inplace=False)
    df.to_excel('combine.xlsx')
    wb_com = openpyxl.load_workbook('combine.xlsx')
    ws_com = wb_com.active
    ws_com.delete_cols(1)
    wb_com.save('Combine.xlsx')
    print("整合完成")
    filename = "combine.xlsx"
    return filename


# 填补空白值
def fill_empty(empty, filename):
    wd = openpyxl.load_workbook(filename)
    ws = wd.active
    for i in empty:
        ws.cell(row=i[0], column=i[1]).value = ws.cell(row=i[0] - 1, column=i[1]).value
    wd.save(filename)
    print("填补空白值成功")


def run(filename1, filename2):
    new_filename2 = 'new_data2.xlsx'
    txt2xlsx(filename2, new_filename2)
    standard(filename1, new_filename2)
    com_filename = combine(filename1, new_filename2)
    empty_pot = remove(com_filename)
    fill_empty(empty_pot, com_filename)


if __name__ == '__main__':
    run('data_1.xlsx', 'data_2.txt')
